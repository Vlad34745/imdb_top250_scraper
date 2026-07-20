import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.parser import fetch_page_html, parse_movies

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

IMDB_URL = "https://www.imdb.com/chart/top/"
OUTPUT_DIR = "output"


def save_to_excel(movies: list, filename: str) -> None:
    df = pd.DataFrame(movies)
    df.columns = ["Rank", "Title", "Year", "Rating", "IMDb ID"]

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = os.path.join(OUTPUT_DIR, filename)

    df.to_excel(filepath, index=False, sheet_name="IMDb Top 250")

    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filepath)
    logger.info(f"Saved {len(movies)} movies to {filepath}")


def main():
    html = fetch_page_html(IMDB_URL)
    movies = parse_movies(html, limit=250)

    if not movies:
        logger.warning("No movies were parsed. Nothing to save.")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"imdb_top250_{timestamp}.xlsx"
    save_to_excel(movies, filename)


if __name__ == "__main__":
    main()